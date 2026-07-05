"""Excel（.xlsx）出力（設計書 6章・視認性改良版）

- 候補一覧: 追記型・URL基準で重複スキップ・数式/リンク/プルダウン付き
  - 判定・利益など意思決定に使う列を左側に集約
  - 原価内訳（商品価格〜販売手数料）は折りたたみグループ（＋ボタンで展開）
  - 判定の色分け / 赤字利益は赤文字 / 縞模様 / オートフィルタ / 枠固定
- 設定: 送料・資材代・判定基準テーブル（数式が参照）
- 仕入れ済み: 在庫管理の入口（空シート）
- 検索リンク: Tier3サイトの検索URL
"""
from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from core.report import CandidateRow
from core.shipping import SIZE_LABELS, ShippingTable
from output import layout

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
LINK_FONT = Font(color="0563C1", underline="single")
ZEBRA_FILL = PatternFill("solid", fgColor="F4F7FB")
THIN_BORDER = Border(bottom=Side(style="hair", color="D0D7DE"))

JUDGE_FILLS = {
    "◎": PatternFill("solid", fgColor="C6EFCE"),
    "○": PatternFill("solid", fgColor="E2EFDA"),
    "△": PatternFill("solid", fgColor="FFF2CC"),
    "×": PatternFill("solid", fgColor="D9D9D9"),
    "要目視": PatternFill("solid", fgColor="DDEBF7"),
}

COLUMN_WIDTHS = [7, 50, 10, 7, 13, 13, 10, 9, 12, 11, 16, 9, 13, 22,
                 10, 10, 10, 10, 9, 8, 10, 8, 15, 15, 20]

MONEY_COLS = "EFGOPQRSTU"
PERCENT_COLS = "HV"


class ExcelWriter:
    def __init__(self, config: dict, path: str | Path):
        self.config = config
        self.path = Path(path)

    # ------------------------------------------------------------------
    def write(self, rows: list[CandidateRow], tier3_links: list[dict]) -> Path:
        wb = self._load_or_create()
        ws = wb[layout.SHEET_CANDIDATES]
        existing_urls = self._existing_urls(ws)

        appended = 0
        for row in rows:
            if row.product.url in existing_urls:
                continue
            self._append_row(ws, row)
            existing_urls.add(row.product.url)
            appended += 1

        self._apply_validations(ws)
        self._apply_conditional_formatting(ws)
        ws.auto_filter.ref = f"A1:Y{max(ws.max_row, 2)}"

        if tier3_links:
            self._append_links(wb[layout.SHEET_LINKS], tier3_links)

        wb.save(self.path)
        logger.info("Excel出力: %s（新規 %d件 / 重複スキップ %d件）",
                    self.path, appended, len(rows) - appended)
        return self.path

    # ------------------------------------------------------------------
    def _load_or_create(self) -> Workbook:
        if not self.path.exists():
            return self._create_workbook()
        wb = load_workbook(self.path)
        # 旧レイアウト（A1が「取得日時」等）のファイルは退避して新規作成
        if (layout.SHEET_CANDIDATES not in wb.sheetnames
                or wb[layout.SHEET_CANDIDATES]["A1"].value != layout.HEADERS[0]):
            backup = self.path.with_name(
                f"{self.path.stem}_旧レイアウト_{datetime.now():%Y%m%d%H%M%S}{self.path.suffix}")
            self.path.rename(backup)
            logger.info("旧レイアウトのファイルを退避しました: %s", backup.name)
            return self._create_workbook()
        return wb

    def _create_workbook(self) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = layout.SHEET_CANDIDATES
        self._write_header(ws, layout.HEADERS)
        ws.freeze_panes = "C2"  # 判定・商品名を固定してスクロール

        for i, w in enumerate(COLUMN_WIDTHS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        # 原価内訳（O〜U）は折りたたみグループに（列番号の左上「＋」で展開）
        start, end = layout.DETAIL_GROUP
        ws.column_dimensions.group(start, end, hidden=True, outline_level=1)

        # 印刷/PDF共有時は意思決定列（A〜N）を横向き・幅合わせで1ページ幅に収める
        ws.print_area = "A1:N10000"
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.print_title_rows = "1:1"  # 各ページにヘッダー行を繰り返す

        settings = wb.create_sheet(layout.SHEET_SETTINGS)
        table = ShippingTable(self.config)
        for row in layout.settings_rows(self.config, table.entries()):
            settings.append(row)
        settings["B1"].number_format = "0%"
        settings["B2"].number_format = "0%"
        for r in (4, 5, 6):
            settings.cell(row=r, column=7).number_format = "0%"
        for col, w in zip("ABCDEFG", [24, 10, 10, 4, 14, 10, 10]):
            settings.column_dimensions[col].width = w

        purchased = wb.create_sheet(layout.SHEET_PURCHASED)
        self._write_header(purchased, layout.HEADERS)

        links = wb.create_sheet(layout.SHEET_LINKS)
        self._write_header(links, layout.LINKS_HEADERS)
        for col, w in zip("ABCD", [17, 30, 16, 60]):
            links.column_dimensions[col].width = w

        return wb

    @staticmethod
    def _write_header(ws, headers: list[str]) -> None:
        ws.append(headers)
        ws.row_dimensions[1].height = 22
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

    @staticmethod
    def _existing_urls(ws) -> set[str]:
        urls: set[str] = set()
        for row in ws.iter_rows(min_row=2, min_col=10, max_col=10):  # J: 商品URL
            cell = row[0]
            if cell.hyperlink and cell.hyperlink.target:
                urls.add(cell.hyperlink.target)
            elif isinstance(cell.value, str) and cell.value.startswith("http"):
                urls.add(cell.value)
        return urls

    # ------------------------------------------------------------------
    def _append_row(self, ws, row: CandidateRow) -> None:
        p = row.product
        r = ws.max_row + 1
        est = row.estimate

        values = {
            "B": p.title,
            "C": p.source,
            "D": {"new": "新品", "used": "中古", "refurbished": "整備済"}.get(p.condition, "不明"),
            "F": est.estimated_price if est.estimated_price else "",
            "K": SIZE_LABELS.get(row.size_key, ""),
            "L": "",
            "M": est.label,
            "N": " / ".join(row.flags),
            "O": p.price,
            "P": p.shipping_cost,
            "Q": p.points,
            "R": p.import_cost,
            "W": p.jan_code or "",
            "X": p.fetched_at.strftime("%m/%d %H:%M"),
            "Y": f"換算: {p.currency_note}" if p.currency_note else "",
        }
        for col, value in values.items():
            ws[f"{col}{r}"] = value
        for col, formula in layout.row_formulas(r).items():
            ws[f"{col}{r}"] = formula

        # ハイパーリンク（I: メルカリ相場確認 / J: 商品ページ）
        if row.mercari_url:
            i = ws[f"I{r}"]
            i.value = "相場を見る"
            i.hyperlink = row.mercari_url
            i.font = LINK_FONT
        j = ws[f"J{r}"]
        j.value = "商品を開く"
        j.hyperlink = p.url
        j.font = LINK_FONT

        for col in MONEY_COLS:
            ws[f"{col}{r}"].number_format = "#,##0"
        for col in PERCENT_COLS:
            ws[f"{col}{r}"].number_format = "0.0%"
        ws[f"A{r}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"A{r}"].font = Font(bold=True, size=12)
        # 商品名は折り返して隣列へのはみ出しを防ぐ
        ws[f"B{r}"].alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[r].height = 30
        for col in "DLM":
            ws[f"{col}{r}"].alignment = Alignment(horizontal="center", vertical="center")
        for c in range(1, len(layout.HEADERS) + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER

    def _apply_validations(self, ws) -> None:
        ws.data_validations.dataValidation = []
        size_dv = DataValidation(
            type="list",
            formula1='"' + ",".join(layout.SIZE_DROPDOWN) + '"',
            allow_blank=True,
            showDropDown=False,
        )
        size_dv.add("K2:K10000")
        ws.add_data_validation(size_dv)

        confirm_dv = DataValidation(type="list", formula1='"✔"', allow_blank=True,
                                    showDropDown=False)
        confirm_dv.add("L2:L10000")
        ws.add_data_validation(confirm_dv)

    def _apply_conditional_formatting(self, ws) -> None:
        ws.conditional_formatting = type(ws.conditional_formatting)()
        # 判定の色分け（A列）
        for mark, fill in JUDGE_FILLS.items():
            ws.conditional_formatting.add(
                "A2:A10000",
                CellIsRule(operator="equal", formula=[f'"{mark}"'], fill=fill),
            )
        # 利益額: 赤字は赤太字、黒字は緑
        ws.conditional_formatting.add(
            "G2:G10000",
            CellIsRule(operator="lessThan", formula=["0"],
                       font=Font(color="CC0000", bold=True)),
        )
        ws.conditional_formatting.add(
            "G2:G10000",
            CellIsRule(operator="greaterThan", formula=["0"],
                       font=Font(color="1A7F37", bold=True)),
        )
        # 縞模様（判定色より低優先。B列以降に適用してA列の判定色を保つ）
        ws.conditional_formatting.add(
            "B2:Y10000",
            FormulaRule(formula=["MOD(ROW(),2)=0"], fill=ZEBRA_FILL),
        )

    # ------------------------------------------------------------------
    def _append_links(self, ws, tier3_links: list[dict]) -> None:
        existing = {
            (row[1].value, row[2].value)
            for row in ws.iter_rows(min_row=2, max_col=4)
        }
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        for link in tier3_links:
            if (link["query"], link["site"]) in existing:
                continue
            r = ws.max_row + 1
            ws[f"A{r}"] = now
            ws[f"B{r}"] = link["query"]
            ws[f"C{r}"] = link["site"]
            cell = ws[f"D{r}"]
            cell.value = link["url"]
            cell.hyperlink = link["url"]
            cell.font = LINK_FONT
