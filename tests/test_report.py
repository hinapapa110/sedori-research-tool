"""候補行組み立て（report.py）と Excel 出力の統合テスト"""
from pathlib import Path

from core.report import build_rows, dedupe_products, filter_rows
from tests.test_profit import make_product

JAN = "4901234567890"


def make_dataset():
    """駿河屋の中古品3,000円、楽天/Yahoo!の中古相場中央値6,000円のデータセット"""
    candidate = make_product(
        source="surugaya", price=3000, jan_code=JAN, condition="used",
        url="https://example.com/candidate", title="人気ゲームソフト",
    )
    refs = [
        make_product(source="yahoo", price=p, jan_code=JAN, condition="used",
                     url=f"https://example.com/r{i}")
        for i, p in enumerate([5500, 6000, 6500])
    ]
    return [candidate] + refs


class TestBuildRows:
    def test_candidate_judged_profitable(self, config):
        rows = build_rows(make_dataset(), config)
        row = next(r for r in rows if r.product.source == "surugaya")
        # 推定販売価格 = 6000 × 0.85 = 5100
        assert row.estimate.estimated_price == 5100
        assert row.estimate.status == "推定OK"
        # 利益 = 5100 - 3000 - 510 - 750 - 80 = 760 / 利益率 14.9% → △
        assert row.profit.profit == 760
        assert row.judgement == "△"
        assert "status=sold_out" in row.mercari_url

    def test_rows_sorted_by_judgement(self, config):
        rows = build_rows(make_dataset(), config)
        orders = ["◎○△".index(r.judgement) if r.judgement in "◎○△" else 9 for r in rows]
        assert orders == sorted(orders)

    def test_skip_rejected_filter(self, config):
        config["output"]["skip_rejected"] = True
        rows = filter_rows(build_rows(make_dataset(), config), config)
        assert all(r.judgement != "×" for r in rows)


class TestDedupe:
    def test_url_based(self):
        a = make_product(url="https://example.com/same")
        b = make_product(url="https://example.com/same", price=9999)
        c = make_product(url="https://example.com/other")
        assert len(dedupe_products([a, b, c])) == 2


class TestExcelOutput:
    def test_write_and_append(self, config, tmp_path: Path):
        from openpyxl import load_workbook

        from output.excel import ExcelWriter

        rows = build_rows(make_dataset(), config)
        links = [{"site": "ヤフオク", "site_key": "yahoo_auction",
                  "query": "人気ゲームソフト", "url": "https://auctions.yahoo.co.jp/search/search?p=x"}]
        path = tmp_path / "out.xlsx"
        writer = ExcelWriter(config, path)
        writer.write(rows, links)

        wb = load_workbook(path)
        assert set(wb.sheetnames) == {"候補一覧", "設定", "仕入れ済み", "検索リンク"}
        ws = wb["候補一覧"]
        assert ws.max_row == 1 + len(rows)
        # 新レイアウト: A列=判定(数式)、E列=実質仕入れ(数式)
        assert ws["A1"].value == "判定"
        assert "要目視" in str(ws["A2"].value)
        assert str(ws["E2"].value).startswith("=IF")
        # 設定シートに手数料率
        assert wb["設定"]["B1"].value == 0.10

        # 同じデータを再出力 → URL重複でスキップされ行数が変わらない
        writer2 = ExcelWriter(config, path)
        writer2.write(rows, links)
        wb2 = load_workbook(path)
        assert wb2["候補一覧"].max_row == ws.max_row
        assert wb2["検索リンク"].max_row == 2  # リンクも重複スキップ
